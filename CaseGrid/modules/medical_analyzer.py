"""CaseGrid – Medical Record Analyzer module."""
from __future__ import annotations

import importlib.util
import io
import json
from pathlib import Path
from typing import Any, Callable

import streamlit as st

# ---------------------------------------------------------------------------
# Template storage
# ---------------------------------------------------------------------------

TEMPLATE_DIR = Path(__file__).resolve().parents[1] / "assets" / "medical_templates"
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
SCRIPTS_DIR = Path(__file__).resolve().parents[1] / "scripts"


def _load_script_function(script_name: str, function_name: str) -> Callable[..., Any] | None:
    module_path = SCRIPTS_DIR / script_name
    if not module_path.exists():
        return None

    spec = importlib.util.spec_from_file_location(function_name, module_path)
    if spec is None or spec.loader is None:
        return None

    module = importlib.util.module_from_spec(spec)
    try:
        spec.loader.exec_module(module)
    except Exception:
        return None
    fn = getattr(module, function_name, None)
    if callable(fn):
        return fn
    return None


NORMALIZE_MEDICAL_RECORD = _load_script_function(
    "medical_normalizer.py", "normalize_medical_record"
)
GENERATE_MEMO = _load_script_function("memo_generator.py", "generate_memo")


def _load_templates() -> dict:
    templates: dict = {}
    for file in sorted(TEMPLATE_DIR.glob("*.json")):
        try:
            with open(file, "r", encoding="utf-8") as f:
                templates[file.stem] = json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return templates


def _save_template(name: str, data: dict) -> None:
    with open(TEMPLATE_DIR / f"{name}.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


# ---------------------------------------------------------------------------
# Rule evaluation engine
# ---------------------------------------------------------------------------

def _evaluate_rule(record_value, operator: str, expected: str):
    """Return True / False / None (insufficient data)."""
    if record_value is None:
        return None

    if operator == "=":
        return str(record_value) == expected
    if operator == "!=":
        return str(record_value) != expected
    if operator == "<":
        try:
            return float(record_value) < float(expected)
        except (ValueError, TypeError):
            return None
    if operator == ">":
        try:
            return float(record_value) > float(expected)
        except (ValueError, TypeError):
            return None
    if operator == "contains":
        return expected.lower() in str(record_value).lower()
    if operator == "not contains":
        return expected.lower() not in str(record_value).lower()
    if operator == "exists":
        return record_value not in (None, "", [])
    if operator == "not exists":
        return record_value in (None, "", [])
    return None


def _evaluate_record(record: dict, template: dict) -> tuple[str, list[dict]]:
    results: list[dict] = []
    for rule in template.get("rules", []):
        value = record.get(rule["field"])
        result = _evaluate_rule(value, rule["operator"], rule["value"])
        results.append({"rule": rule, "value": value, "result": result})

    insufficient = any(
        r["result"] is None and r["rule"].get("required", True)
        for r in results
    )
    if insufficient:
        status = "Insufficient Data"
    else:
        qualifies = all(
            r["result"] is True or not r["rule"].get("required", True)
            for r in results
        )
        status = "Qualified" if qualifies else "Not Qualified"

    return status, results


# ---------------------------------------------------------------------------
# XLSX export helper
# ---------------------------------------------------------------------------

def _build_xlsx(results: list[dict]) -> bytes:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis Results"

        STATUS_COLORS = {
            "Qualified": "C6EFCE",
            "Not Qualified": "FFC7CE",
            "Insufficient Data": "FFEB9C",
        }

        # Header row
        headers = ["File", "Status"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True)

        for row_idx, r in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=r["filename"])
            status_cell = ws.cell(row=row_idx, column=2, value=r["status"])
            color = STATUS_COLORS.get(r["status"])
            if color:
                status_cell.fill = PatternFill("solid", fgColor=color)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except ImportError:
        return b""


# ---------------------------------------------------------------------------
# Main Streamlit UI
# ---------------------------------------------------------------------------

FRIENDLY_OPERATORS: dict[str, str] = {
    "is equal to": "=",
    "is not equal to": "!=",
    "is less than": "<",
    "is greater than": ">",
    "contains": "contains",
    "does not contain": "not contains",
    "is present": "exists",
    "is missing": "not exists",
}
_FRIENDLY_LIST = list(FRIENDLY_OPERATORS.keys())
_SK = "med_"  # session-state key prefix


def run() -> None:
    st.header("Medical Record Analyzer")

    if NORMALIZE_MEDICAL_RECORD is None:
        st.error("Missing parser: CaseGrid/scripts/medical_normalizer.py")
        return

    # ---- Load persisted templates ----
    templates = _load_templates()
    template_names = list(templates.keys())

    # ---- Template controls ----
    st.subheader("Template Controls")
    t_col1, t_col2, t_col3, t_col4 = st.columns(4)

    with t_col1:
        selected_template = st.selectbox(
            "Load template",
            ["— select —"] + template_names,
            key=f"{_SK}sel_template",
        )

    with t_col2:
        new_name = st.text_input("Template name", key=f"{_SK}new_name")
        if st.button("New Template", key=f"{_SK}btn_new"):
            st.session_state[f"{_SK}current"] = {
                "name": new_name or "Untitled",
                "rules": [],
            }
            st.rerun()

    with t_col3:
        if st.button("Save Template", key=f"{_SK}btn_save"):
            tpl = st.session_state.get(f"{_SK}current")
            if tpl:
                save_name = tpl.get("name", "Untitled")
                _save_template(save_name, tpl)
                st.success(f"Saved **{save_name}**")
            else:
                st.warning("No active template to save.")

    with t_col4:
        if (
            st.button("Delete Template", key=f"{_SK}btn_delete")
            and selected_template != "— select —"
        ):
            target = TEMPLATE_DIR / f"{selected_template}.json"
            target.unlink(missing_ok=True)
            st.success(f"Deleted **{selected_template}**")
            st.rerun()

    if f"{_SK}current" not in st.session_state:
        st.session_state[f"{_SK}current"] = {"name": "New Template", "rules": []}

    # Load selected template into session state when selection changes.
    loaded_template_key = f"{_SK}loaded_template"
    last_loaded_template = st.session_state.get(loaded_template_key)
    if (
        selected_template != "— select —"
        and selected_template != last_loaded_template
    ):
        st.session_state[f"{_SK}current"] = templates.get(
            selected_template, {"name": selected_template, "rules": []}
        )
        st.session_state[loaded_template_key] = selected_template

    st.divider()

    # ---- Rule builder ----
    st.subheader("Qualification Rules")
    st.caption("Use plain-English conditions. CaseGrid handles the logic automatically.")
    rules: list[dict] = st.session_state[f"{_SK}current"]["rules"]

    for i, rule in enumerate(rules):
        rA, rB, rC, rD, rE = st.columns([2, 2, 2, 1, 1])

        with rA:
            rule["field"] = st.text_input(
                "Field", value=rule.get("field", ""), key=f"{_SK}field_{i}"
            )
        with rB:
            current_friendly = next(
                (k for k, v in FRIENDLY_OPERATORS.items() if v == rule.get("operator")),
                _FRIENDLY_LIST[0],
            )
            selected_friendly = st.selectbox(
                f"Condition {i + 1}",
                _FRIENDLY_LIST,
                index=_FRIENDLY_LIST.index(current_friendly),
                key=f"{_SK}op_{i}",
            )
            rule["operator"] = FRIENDLY_OPERATORS[selected_friendly or _FRIENDLY_LIST[0]]
        with rC:
            rule["value"] = st.text_input(
                "Value", value=rule.get("value", ""), key=f"{_SK}val_{i}"
            )
        with rD:
            rule["required"] = st.checkbox(
                "Required", value=rule.get("required", True), key=f"{_SK}req_{i}"
            )
        with rE:
            st.write("")  # vertical spacer
            if st.button("–", key=f"{_SK}rm_{i}"):
                rules.pop(i)
                st.rerun()

    if st.button("+ Add Rule", key=f"{_SK}add_rule"):
        rules.append({"field": "", "operator": "=", "value": "", "required": True})

    st.divider()

    # ---- File upload + analysis ----
    st.subheader("Upload Medical Records")

    if f"{_SK}uploader_key" not in st.session_state:
        st.session_state[f"{_SK}uploader_key"] = 0

    up_col1, up_col2 = st.columns([3, 1])
    with up_col2:
        if st.button("Clear Files", key=f"{_SK}clear"):
            st.session_state[f"{_SK}uploader_key"] += 1
            st.session_state.pop(f"{_SK}results", None)
            for k in list(st.session_state.keys()):
                if k.startswith(f"{_SK}memo_file_"):
                    st.session_state.pop(k, None)
            st.rerun()

    uploads = st.file_uploader(
        "Upload PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        key=f"{_SK}uploads_{st.session_state[f'{_SK}uploader_key']}",
    )

    if st.button("Run Analysis", key=f"{_SK}run") and uploads:
        if not rules:
            st.warning("Add at least one qualification rule before running.")
        else:
            results: list[dict] = []
            with st.spinner("Analyzing records…"):
                for upload in uploads:
                    record = NORMALIZE_MEDICAL_RECORD(upload.getvalue())
                    status, rule_results = _evaluate_record(
                        record, st.session_state[f"{_SK}current"]
                    )
                    missing_data = [
                        r["rule"].get("field", "")
                        for r in rule_results
                        if r["result"] is None and r["rule"].get("required", True)
                    ]
                    patient_name = Path(upload.name).stem
                    page_count = record.get("page_count") or 0
                    results.append(
                        {
                            "filename": upload.name,
                            "patient_name": patient_name,
                            "record_range": f"Pages 1-{page_count}" if page_count else "Unknown",
                            "status": status,
                            "details": rule_results,
                            "record": record,
                            "missing_data": missing_data,
                        }
                    )
            st.session_state[f"{_SK}results"] = results

    # ---- Results ----
    if f"{_SK}results" in st.session_state:
        results = st.session_state[f"{_SK}results"]
        st.subheader("Results")

        # Summary counts
        counts = {"Qualified": 0, "Not Qualified": 0, "Insufficient Data": 0}
        for r in results:
            counts[r["status"]] = counts.get(r["status"], 0) + 1

        c1, c2, c3 = st.columns(3)
        c1.metric("Qualified", counts["Qualified"])
        c2.metric("Not Qualified", counts["Not Qualified"])
        c3.metric("Insufficient Data", counts["Insufficient Data"])

        # Results table
        table_data = [
            {
                "File": r["filename"],
                "Patient": r["patient_name"],
                "Status": r["status"],
                "Age At Diagnosis": r["record"].get("age_at_diagnosis"),
                "Primary Cancer": r["record"].get("primary_cancer"),
                "BRCA": r["record"].get("brca_result"),
            }
            for r in results
        ]
        st.dataframe(table_data, use_container_width=True)

        # Detailed breakdown per file
        with st.expander("Detailed Rule Breakdown"):
            for r in results:
                st.markdown(f"**{r['filename']}** — _{r['status']}_")
                for rd in r["details"]:
                    rule = rd["rule"]
                    icon = (
                        "✅" if rd["result"] is True
                        else "❌" if rd["result"] is False
                        else "⚠️"
                    )
                    req_label = "(required)" if rule.get("required") else "(optional)"
                    st.write(
                        f"{icon} `{rule['field']}` {rule['operator']} `{rule['value']}` "
                        f"{req_label} — extracted: `{rd['value']}`"
                    )
                st.divider()

        st.subheader("Legal Memo Generator")
        if GENERATE_MEMO is None:
            st.warning("Memo generator unavailable. Add CaseGrid/scripts/memo_generator.py")
        else:
            for idx, r in enumerate(results):
                if st.button(f"Generate Legal Memo: {r['filename']}", key=f"{_SK}memo_{idx}"):
                    memo_path = GENERATE_MEMO(
                        r["patient_name"],
                        r["record"].get("dob"),
                        r["record_range"],
                        {"status": r["status"], "details": r["details"]},
                        r["record"].get("evidence", []),
                        r["missing_data"],
                    )
                    memo_file = Path(memo_path)
                    st.session_state[f"{_SK}memo_file_{idx}"] = str(memo_file)
                    st.success(f"Memo generated: {memo_file}")

                memo_file_str = st.session_state.get(f"{_SK}memo_file_{idx}")
                if memo_file_str:
                    memo_file = Path(memo_file_str)
                    if memo_file.exists():
                        st.download_button(
                            f"Download Memo: {r['patient_name']}",
                            data=memo_file.read_bytes(),
                            file_name=memo_file.name,
                            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                            key=f"{_SK}memo_dl_{idx}",
                        )

        # XLSX download
        xlsx_bytes = _build_xlsx(results)
        if xlsx_bytes:
            st.download_button(
                "Download Results (.xlsx)",
                data=xlsx_bytes,
                file_name="medical_analysis_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{_SK}dl_xlsx",
            )
