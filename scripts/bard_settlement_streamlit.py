#!/usr/bin/env python3
"""Local Streamlit UI for Bard settlement PDF batch processing.

Run:
    streamlit run scripts/bard_settlement_streamlit.py
"""

from __future__ import annotations

import io
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from time import perf_counter

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side

try:
    from scripts.bard_pdf_extract import (
        build_new_pdf_name,
        extract_fields_from_pdf,
        process_pdf_file,
        unique_path,
    )
except ModuleNotFoundError:
    import sys

    # Fallback 1: add repo root and retry package-style import.
    repo_root = Path(__file__).resolve().parents[1]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))

    try:
        from scripts.bard_pdf_extract import (
            build_new_pdf_name,
            extract_fields_from_pdf,
            process_pdf_file,
            unique_path,
        )
    except ModuleNotFoundError:
        # Fallback 2: import sibling module directly from scripts/ directory.
        scripts_dir = Path(__file__).resolve().parent
        if str(scripts_dir) not in sys.path:
            sys.path.insert(0, str(scripts_dir))
        from bard_pdf_extract import (
            build_new_pdf_name,
            extract_fields_from_pdf,
            process_pdf_file,
            unique_path,
        )

INPUT_EXT = ".pdf"


def _split_client_name(client_name: str) -> tuple[str, str]:
    if "," in client_name:
        last, first = client_name.split(",", 1)
        return last.strip(), first.strip()
    parts = client_name.split()
    if len(parts) >= 2:
        return parts[-1], " ".join(parts[:-1])
    return client_name.strip(), ""


def _build_settlements_workbook(rows: list[tuple[str, str, float]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")
    ws.title = "Settlements"
    bold_font = Font(bold=True)
    thin_side = Side(style="thin", color="000000")
    all_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.append(
        [
            "No.",
            "Claim ID:",
            "Claimant Last Name",
            "Claimant First Name",
            "Delivery Method",
            "Scheduled Send Date",
            "Total Award",
        ]
    )
    for col_idx in range(1, 8):
        ws.cell(row=1, column=col_idx).font = bold_font

    sorted_rows = sorted(rows, key=lambda row: _split_client_name(row[1])[0].lower())
    for idx, (claim_number, client_name, award_float) in enumerate(sorted_rows, start=1):
        last_name, first_name = _split_client_name(client_name)
        ws.append([idx, claim_number, last_name, first_name, "", "", award_float])

    if sorted_rows:
        data_start = 2
        data_end = len(sorted_rows) + 1
        for row_idx in range(data_start, data_end + 1):
            ws.cell(row=row_idx, column=7).number_format = "$#,##0.00"

        summary_row = data_end + 1
        ws.cell(row=summary_row, column=4, value="Average Settlement")
        ws.cell(row=summary_row, column=5, value=f"=AVERAGE(G{data_start}:G{data_end})")
        ws.cell(row=summary_row, column=6, value="Total Settlements")
        ws.cell(row=summary_row, column=7, value=f"=SUM(G{data_start}:G{data_end})")
        ws.cell(row=summary_row, column=5).number_format = "$#,##0.00"
        ws.cell(row=summary_row, column=7).number_format = "$#,##0.00"
        for col_idx in range(1, 8):
            ws.cell(row=summary_row, column=col_idx).font = bold_font

    table_end_row = len(sorted_rows) + 2 if sorted_rows else 1
    for row_idx in range(1, table_end_row + 1):
        for col_idx in range(1, 8):
            ws.cell(row=row_idx, column=col_idx).border = all_border

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _preview_rows(uploaded_files) -> list[dict]:
    rows: list[dict] = []
    used_names: set[str] = set()

    for upload in uploaded_files:
        original_name = upload.name
        try:
            if not original_name.lower().endswith(INPUT_EXT):
                rows.append(
                    {
                        "original_name": original_name,
                        "proposed_name": "",
                        "claim_number": "",
                        "client_name": "",
                        "total_award": "",
                        "status": "error",
                        "error": "Unsupported file type (must be .pdf)",
                    }
                )
                continue

            with tempfile.TemporaryDirectory(prefix="bard_preview_") as td:
                tmp_path = Path(td) / original_name
                tmp_path.write_bytes(upload.getvalue())
                claim_number, client_name, award_float, award_formatted = extract_fields_from_pdf(tmp_path)

            if not (claim_number and client_name and award_formatted):
                rows.append(
                    {
                        "original_name": original_name,
                        "proposed_name": "",
                        "claim_number": claim_number or "",
                        "client_name": client_name or "",
                        "total_award": award_formatted or "",
                        "status": "error",
                        "error": "Missing one or more required fields",
                    }
                )
                continue

            proposed = build_new_pdf_name(client_name, claim_number, award_formatted)
            stem = Path(proposed).stem
            suffix = Path(proposed).suffix
            candidate = proposed
            idx = 1
            while candidate in used_names:
                candidate = f"{stem}_{idx}{suffix}"
                idx += 1
            used_names.add(candidate)

            rows.append(
                {
                    "original_name": original_name,
                    "proposed_name": candidate,
                    "claim_number": claim_number,
                    "client_name": client_name,
                    "total_award": award_formatted,
                    "status": "success",
                    "error": "",
                }
            )
        except Exception as ex:
            rows.append(
                {
                    "original_name": original_name,
                    "proposed_name": "",
                    "claim_number": "",
                    "client_name": "",
                    "total_award": "",
                    "status": "error",
                    "error": str(ex),
                }
            )

    return rows


def _build_download(uploaded_files) -> tuple[bytes, str, str, bool, int, int]:
    with tempfile.TemporaryDirectory(prefix="bard_pdf_streamlit_") as td:
        root = Path(td)
        input_dir = root / "input"
        output_dir = root / "output"
        package_dir = root / "package"

        input_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)
        package_dir.mkdir(parents=True, exist_ok=True)

        results: list[dict] = []
        settlement_rows: list[tuple[str, str, float]] = []

        for upload in uploaded_files:
            original_name = Path(upload.name).name
            try:
                if not original_name.lower().endswith(INPUT_EXT):
                    raise ValueError("Unsupported file type (must be .pdf)")

                src_path = unique_path(input_dir / original_name)
                src_path.write_bytes(upload.getvalue())

                new_path, meta = process_pdf_file(src_path, package_dir)

                settlement_rows.append((meta["claim_number"], meta["client_name"], meta["award_float"]))

                results.append(
                    {
                        "original_name": original_name,
                        "renamed_pdf": new_path.name,
                        "status": "success",
                    }
                )
            except Exception as ex:
                results.append(
                    {
                        "original_name": original_name,
                        "status": "error",
                        "error": str(ex),
                    }
                )

        # Write XLSX summary alongside the PDFs
        if settlement_rows:
            xlsx_path = package_dir / "settlements_summary.xlsx"
            xlsx_path.write_bytes(_build_settlements_workbook(settlement_rows))

        n_ok = sum(1 for r in results if r["status"] == "success")
        n_err = sum(1 for r in results if r["status"] == "error")

        packaged_items = sorted(item for item in package_dir.glob("*") if item.is_file())
        if len(packaged_items) == 1:
            single_item = packaged_items[0]
            suffix = single_item.suffix.lower()
            mime = (
                "application/pdf"
                if suffix == ".pdf"
                else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            return (
                single_item.read_bytes(),
                single_item.name,
                mime,
                False,
                n_ok,
                n_err,
            )

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_name = f"processed_bard_settlements_{ts}.zip"
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for item in packaged_items:
                zf.write(item, arcname=item.name)

        return buf.getvalue(), zip_name, "application/zip", True, n_ok, n_err


def main() -> None:
    st.set_page_config(page_title="Bard Settlement PDF Processor", layout="wide")

    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0

    st.title("Bard Settlement PDF Processor")
    st.caption(
        "Upload Bard settlement PDFs, preview extracted fields and rename output, "
        "and download processed files plus Excel summary."
    )

    uploaded_files = st.file_uploader(
        "Upload settlement PDF files",
        type=["pdf"],
        accept_multiple_files=True,
        key=f"uploader_{st.session_state['uploader_key']}",
    )

    col1, col2, col3 = st.columns(3)

    with col3:
        if st.button("Clear Files"):
            st.session_state["uploader_key"] += 1
            st.session_state.pop("preview_rows", None)
            st.session_state.pop("download_bytes", None)
            st.session_state.pop("download_name", None)
            st.session_state.pop("download_mime", None)
            st.session_state.pop("is_zip", None)
            st.session_state.pop("process_summary", None)
            st.session_state.pop("process_timing", None)
            st.rerun()

    with col1:
        if st.button("Preview Renames", disabled=not uploaded_files):
            rows = _preview_rows(uploaded_files)
            st.session_state["preview_rows"] = rows

    with col2:
        if st.button("Process Batch", type="primary", disabled=not uploaded_files):
            with st.spinner("Processing uploaded files..."):
                files_to_process = uploaded_files or []
                start = perf_counter()
                download_bytes, download_name, download_mime, is_zip, n_ok, n_err = _build_download(
                    files_to_process
                )
                elapsed_seconds = perf_counter() - start
                processed_count = n_ok + n_err
                total_input_bytes = sum(
                    int(getattr(upload, "size", 0) or len(upload.getvalue()))
                    for upload in files_to_process
                )
                st.session_state["download_bytes"] = download_bytes
                st.session_state["download_name"] = download_name
                st.session_state["download_mime"] = download_mime
                st.session_state["is_zip"] = is_zip
                st.session_state["process_summary"] = (n_ok, n_err)
                st.session_state["process_timing"] = (elapsed_seconds, processed_count, total_input_bytes)

    preview_rows = st.session_state.get("preview_rows")
    if preview_rows:
        st.subheader("Preview Results")
        st.dataframe(preview_rows, use_container_width=True)

    download_bytes = st.session_state.get("download_bytes")
    download_name = st.session_state.get("download_name", "processed_output.zip")
    download_mime = st.session_state.get("download_mime", "application/zip")
    is_zip = st.session_state.get("is_zip", True)
    process_summary = st.session_state.get("process_summary")
    process_timing = st.session_state.get("process_timing")

    if process_summary:
        n_ok, n_err = process_summary
        if n_err:
            st.warning(f"Processed {n_ok} file(s) successfully. {n_err} failed.")
        else:
            st.success(f"Processed {n_ok} file(s) successfully.")

    if process_timing:
        elapsed_seconds, processed_count, total_input_bytes = process_timing
        rate = processed_count / elapsed_seconds if elapsed_seconds > 0 else 0.0
        total_input_mb = total_input_bytes / (1024 * 1024)
        st.info(
            f"Processing time: {elapsed_seconds:.2f}s for {processed_count} file(s) "
            f"({rate:.2f} files/sec). Total input size: {total_input_mb:.2f} MB."
        )

    if download_bytes:
        label_prefix = "Download ZIP" if is_zip else "Download File"
        st.download_button(
            label=f"{label_prefix} ({download_name})",
            data=download_bytes,
            file_name=download_name,
            mime=download_mime,
        )


if __name__ == "__main__":
    main()
